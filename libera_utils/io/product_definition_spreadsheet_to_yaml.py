"""Convert a tab of the Libera Data Product Specifications workbook into a product definition YAML file.

The Libera Data Product Specifications workbook is authored and maintained by the L2 algorithm
developers. Each tab describes one data product: a free-form header block, then three tables
introduced by marker rows in column A (product-specific file-level metadata, coordinates, and
variables). This module turns one of those tabs into the ``attributes`` / ``coordinates`` /
``variables`` structure consumed by
:class:`~libera_utils.io.product_definition.LiberaDataProductDefinition`.

Two design rules govern everything in this module.

**The workbook is strictly read-only.** The tool opens it, reads it, and closes it. It never calls
``Workbook.save``, never rewrites a cell, and never emits a "corrected" copy of the spreadsheet.
The workbook is being actively authored by the algorithm developers and it is theirs. When this
module finds a problem, its only job is to describe the problem well enough that a developer can
fix the spreadsheet by hand and re-run. Please do not add an autofix feature here.

**Fail loudly, but only once.** Anything unexpected is an error rather than a warning, and no
output file is written unless the resulting definition validates cleanly against the pydantic
model. Problems are accumulated across the whole sheet and reported together, with A1-style cell
references, so a developer can fix ten cells in one pass instead of ten run-fix cycles.

This module deliberately does *not* read, load, or reason about
``libera_utils/data/required_product_attributes.yml``. The existing product definition code
already populates and enforces the required standard attributes; conflicts surface during the
round-trip validation below and are reported through the same message formatter as everything
else.
"""

import argparse
import ast
import difflib
import logging
import math
import re
import textwrap
from dataclasses import dataclass, field
from datetime import UTC, datetime
from enum import Enum
from typing import Any, cast

import openpyxl
import yaml
from cloudpathlib import AnyPath
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from libera_utils.constants import DataProductIdentifier
from libera_utils.io.filenaming import PathType
from libera_utils.io.product_definition import LiberaDataProductDefinition
from libera_utils.logutil import configure_task_logging
from libera_utils.version import version as libera_utils_version

logger = logging.getLogger(__name__)

# The cell on every tab holding the "Algorithm:" value. It must agree with the tab name.
ALGORITHM_CELL = "B1"

# Matches a cell whose text is NaN in any casing, with or without a sign.
_NAN_PATTERN = re.compile(r"^[+-]?nan$", re.IGNORECASE)

# Width used when wrapping problem text in rendered error messages.
_MESSAGE_WIDTH = 96


class _Sentinel(Enum):
    """Distinguishes the two kinds of "no value" a spreadsheet cell can hold.

    A blank cell and a cell reading ``N/A`` mean different things. In the file-level metadata
    table a blank means "this attribute is calculated by the algorithm", which becomes an explicit
    null in the YAML, while ``N/A`` means the attribute does not apply and is omitted entirely.
    Collapsing both to ``None`` would lose that distinction.
    """

    EMPTY = "EMPTY"
    NOT_APPLICABLE = "NOT_APPLICABLE"


EMPTY = _Sentinel.EMPTY
NOT_APPLICABLE = _Sentinel.NOT_APPLICABLE

# Section key -> (marker text in column A, header text that must appear in column A one row below).
# A column A cell counts as a section marker only if the row immediately beneath it holds the
# matching header text. That anchoring is what lets a lowercase "variables" be a legitimate marker
# on one tab and an illegal banner row on another.
SPREADSHEET_MARKERS: dict[str, tuple[str, str]] = {
    "attributes": ("Product-Specific File-Level Metadata", "Attribute"),
    "coordinates": ("Coordinates", "Coordinate Name"),
    "variables": ("Variables", "Field Name"),
}

# Words that may not be used as banner text inside a table, because they are section markers.
# Derived from SPREADSHEET_MARKERS so the two cannot drift apart.
RESERVED_BANNER_WORDS: frozenset[str] = frozenset(
    [key.lower() for key in SPREADSHEET_MARKERS] + [marker.lower() for marker, _ in SPREADSHEET_MARKERS.values()]
)


class ColumnKind(Enum):
    """How the value in a column is interpreted after generic normalization.

    Attributes
    ----------
    TEXT
        Never coerced to a number and never parsed as a literal. Anything non-empty becomes a
        string. This is what keeps ``value_meaning`` a string even when the cell reads ``0``.
    SCALAR
        Numeric-looking text is coerced to ``int`` or ``float``; anything else stays a string.
    SEQUENCE
        Must parse to a list. Anything that does not is an error rather than a bare string.
    """

    TEXT = "TEXT"
    SCALAR = "SCALAR"
    SEQUENCE = "SEQUENCE"


class ColumnTarget(Enum):
    """Where a column's value lands in the resulting variable definition."""

    NAME = "NAME"
    TOP_LEVEL = "TOP_LEVEL"
    ATTRIBUTE = "ATTRIBUTE"
    ENCODING = "ENCODING"
    IGNORE = "IGNORE"


@dataclass(frozen=True)
class ColumnSpec:
    """Describes one recognized spreadsheet column.

    Attributes
    ----------
    key: str
        The key this column contributes to the output, e.g. ``long_name`` or ``dtype``. Empty for
        name and ignored columns.
    target: ColumnTarget
        Where the value lands in the variable definition.
    kind: ColumnKind
        How the raw cell value is interpreted.
    collapse_newlines: bool
        Whether internal newlines collapse to single spaces. True only for prose that the
        spreadsheet wraps for readability; ``value_meaning`` uses one line per quality bit and
        that structure is meaningful, so it is preserved.
    """

    key: str
    target: ColumnTarget
    kind: ColumnKind
    collapse_newlines: bool = False


# Recognized columns in the Coordinates and Variables tables, keyed by normalized header text.
# Any header not found here (and not an "encoding.*" column) is an error, so that adding a column
# to the workbook forces a deliberate code update rather than silently dropping the new field.
VARIABLE_COLUMNS: dict[str, ColumnSpec] = {
    "coordinate name": ColumnSpec("", ColumnTarget.NAME, ColumnKind.TEXT),
    "field name": ColumnSpec("", ColumnTarget.NAME, ColumnKind.TEXT),
    "long_name": ColumnSpec("long_name", ColumnTarget.ATTRIBUTE, ColumnKind.TEXT, collapse_newlines=True),
    "dimensions": ColumnSpec("dimensions", ColumnTarget.TOP_LEVEL, ColumnKind.SEQUENCE),
    "dtype": ColumnSpec("dtype", ColumnTarget.TOP_LEVEL, ColumnKind.TEXT),
    "units": ColumnSpec("units", ColumnTarget.ATTRIBUTE, ColumnKind.TEXT),
    "valid_range": ColumnSpec("valid_range", ColumnTarget.ATTRIBUTE, ColumnKind.SEQUENCE),
    "_fillvalue": ColumnSpec("_FillValue", ColumnTarget.ATTRIBUTE, ColumnKind.SCALAR),
    "value_meaning (only applicable to bit flags)": ColumnSpec(
        "value_meaning", ColumnTarget.ATTRIBUTE, ColumnKind.TEXT
    ),
    "value_meaning": ColumnSpec("value_meaning", ColumnTarget.ATTRIBUTE, ColumnKind.TEXT),
    "used/read by": ColumnSpec("", ColumnTarget.IGNORE, ColumnKind.TEXT),
    "notes": ColumnSpec("", ColumnTarget.IGNORE, ColumnKind.TEXT),
}

# Recognized columns in the product-specific file-level metadata table.
METADATA_COLUMNS: dict[str, ColumnSpec] = {
    "attribute": ColumnSpec("", ColumnTarget.NAME, ColumnKind.TEXT),
    "value (leave blank if calculated by the algorithm)": ColumnSpec("", ColumnTarget.ATTRIBUTE, ColumnKind.SCALAR),
    "value": ColumnSpec("", ColumnTarget.ATTRIBUTE, ColumnKind.SCALAR),
    "note": ColumnSpec("", ColumnTarget.IGNORE, ColumnKind.TEXT),
}

# Prefix for dotted headers that nest under a variable's "encoding" key. The nesting is generic on
# the ".", so a new "encoding.<something>" column works without a code change here.
ENCODING_PREFIX = "encoding."

# Attributes this module injects into every product definition, in this order, before the rows of
# the file-level metadata table.
INJECTED_ATTRIBUTES: tuple[str, ...] = ("ProductID", "algorithm_version")

# Top-level key order in the emitted YAML.
DEFINITION_KEY_ORDER: tuple[str, ...] = ("attributes", "coordinates", "variables")


class ProductDefinitionSpreadsheetError(Exception):
    """Raised when a product definition cannot be built from a spreadsheet."""


class ProductIdNotFoundError(ProductDefinitionSpreadsheetError):
    """Raised when the requested data product has no tab in the workbook."""


class MalformedSheetError(ProductDefinitionSpreadsheetError):
    """Raised when a sheet's structure or contents cannot be interpreted."""


@dataclass(frozen=True)
class SpreadsheetProblem:
    """One problem found in a spreadsheet, located at a specific cell.

    Attributes
    ----------
    row: int
        1-based row index of the offending cell.
    column: int
        1-based column index of the offending cell.
    cell_reference: str
        A1-style reference including the sheet name, e.g. ``RAD-4CH!A38``.
    message: str
        What is wrong and, where derivable, what to do about it.
    """

    row: int
    column: int
    cell_reference: str
    message: str


@dataclass
class ParsedSheet:
    """The result of parsing one spreadsheet tab.

    Attributes
    ----------
    definition: dict[str, Any]
        The product definition, with top-level keys ``attributes``, ``coordinates``, and
        ``variables``, ready to be validated and dumped to YAML.
    banners: dict[str, str]
        Maps a variable name to the banner text that preceded it in the Variables table, e.g.
        ``{"Latitude": "Geolocation"}``. Purely cosmetic; used to reproduce the ``# Geolocation #``
        comment lines in the emitted YAML.
    sheet_name: str
        The tab the definition came from.
    """

    definition: dict[str, Any]
    banners: dict[str, str] = field(default_factory=dict)
    sheet_name: str = ""


@dataclass
class _ProblemLog:
    """Accumulates problems found while parsing one sheet, then raises them all at once.

    Attributes
    ----------
    sheet_name: str
        The tab being parsed, used to build A1-style cell references.
    problems: list[SpreadsheetProblem]
        Problems found so far, in discovery order.
    """

    sheet_name: str
    problems: list[SpreadsheetProblem] = field(default_factory=list)

    @property
    def quoted_sheet_name(self) -> str:
        """Sheet name quoted the way a spreadsheet reference requires, if it needs quoting."""
        if re.fullmatch(r"[A-Za-z0-9_.\-]+", self.sheet_name):
            return self.sheet_name
        return f"'{self.sheet_name}'"

    def cell_reference(self, row: int, column: int) -> str:
        """Build an A1-style reference including the sheet name.

        Parameters
        ----------
        row : int
            1-based row index.
        column : int
            1-based column index.

        Returns
        -------
        str
            A reference such as ``RAD-4CH!A38``.
        """
        return f"{self.quoted_sheet_name}!{get_column_letter(column)}{row}"

    def add(self, row: int, column: int, message: str) -> None:
        """Record a problem at a cell.

        Parameters
        ----------
        row : int
            1-based row index of the offending cell.
        column : int
            1-based column index of the offending cell.
        message : str
            What is wrong and what to do about it.
        """
        self.problems.append(SpreadsheetProblem(row, column, self.cell_reference(row, column), message))

    def __bool__(self) -> bool:
        """True when at least one problem has been recorded."""
        return bool(self.problems)

    def render(self, product_id: DataProductIdentifier, source_name: str) -> str:
        """Render every recorded problem as a single user-facing message.

        Parameters
        ----------
        product_id : DataProductIdentifier
            The product the caller asked for.
        source_name : str
            Display name of the source workbook, normally its filename.

        Returns
        -------
        str
            The full multi-line message, ending with confirmation that nothing was written.
        """
        ordered = sorted(self.problems, key=lambda problem: (problem.row, problem.column))
        count = len(ordered)
        plural = "problem" if count == 1 else "problems"
        lines = [
            f"Cannot build a product definition for {product_id} from sheet "
            f"'{self.sheet_name}' of {source_name}. Found {count} {plural}:",
            "",
        ]
        for problem in ordered:
            prefix = f"  {problem.cell_reference}  "
            wrapped = textwrap.wrap(problem.message, width=_MESSAGE_WIDTH) or [""]
            lines.append(prefix + wrapped[0])
            lines.extend(" " * len(prefix) + continuation for continuation in wrapped[1:])
            lines.append("")
        lines.append("Correct these in the spreadsheet and re-run. No output file was written.")
        return "\n".join(lines)

    def raise_if_any(self, product_id: DataProductIdentifier, source_name: str) -> None:
        """Raise :class:`MalformedSheetError` if any problems were recorded.

        Parameters
        ----------
        product_id : DataProductIdentifier
            The product the caller asked for.
        source_name : str
            Display name of the source workbook.

        Raises
        ------
        MalformedSheetError
            If at least one problem was recorded.
        """
        if self.problems:
            raise MalformedSheetError(self.render(product_id, source_name))


def _normalize_marker(text: Any) -> str:
    """Normalize column A text for marker and banner comparisons.

    Lowercases, replaces non-breaking spaces, and collapses runs of whitespace, so that
    ``"Variables"``, ``"variables"``, and ``"Variables\xa0"`` all compare equal.

    Parameters
    ----------
    text : Any
        Raw cell value.

    Returns
    -------
    str
        The normalized comparison key, empty for blank cells.
    """
    if not isinstance(text, str):
        return "" if text is None else str(text).strip().lower()
    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip().lower()


def _is_nan(value: Any) -> bool:
    """Determine whether a value is NaN, in either float or text form.

    Parameters
    ----------
    value : Any
        A normalized cell value.

    Returns
    -------
    bool
        True if the value is a float NaN or a string spelling of NaN.
    """
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and bool(_NAN_PATTERN.match(value.strip()))


def _normalize_cell(raw: Any) -> Any:
    """Apply the generic normalization every cell gets before column-specific handling.

    Blank, whitespace-only, and non-breaking-space-only cells become :data:`EMPTY`. Cells reading
    ``N/A`` in any casing become :data:`NOT_APPLICABLE`. Everything else is stripped of surrounding
    whitespace and returned as-is; numeric coercion and sequence parsing are deliberately *not*
    done here, because they are only correct for specific columns.

    Parameters
    ----------
    raw : Any
        The raw value openpyxl returned for the cell.

    Returns
    -------
    Any
        :data:`EMPTY`, :data:`NOT_APPLICABLE`, or the cleaned value.
    """
    if raw is None:
        return EMPTY
    if isinstance(raw, str):
        cleaned = raw.replace("\xa0", " ").strip()
        if not cleaned:
            return EMPTY
        if cleaned.lower() in {"n/a", "na", "not applicable"}:
            return NOT_APPLICABLE
        return cleaned
    return raw


def _try_literal_eval(text: str) -> Any:
    """Evaluate a Python literal, returning None if it is not one.

    Uses :func:`ast.literal_eval` rather than ``eval`` so there is nothing here for a security
    scanner to flag. Returns None rather than swallowing the exception inline, which keeps the
    caller free of bare ``except: continue`` blocks.

    Parameters
    ----------
    text : str
        Candidate literal text.

    Returns
    -------
    Any
        The evaluated literal, or None if the text is not a Python literal.
    """
    try:
        return ast.literal_eval(text)
    except (ValueError, SyntaxError, MemoryError, RecursionError):
        return None


def _try_yaml_load(text: str) -> Any:
    """Load a YAML scalar or sequence, returning None if the text is not valid YAML.

    Parameters
    ----------
    text : str
        Candidate YAML text.

    Returns
    -------
    Any
        The loaded object, or None if the text does not parse.
    """
    try:
        return yaml.safe_load(text)
    except yaml.YAMLError:
        return None


def _parse_sequence(value: Any) -> list | None:
    """Parse a cell value that is declared to hold a sequence.

    Handles the two literal styles the workbook uses, a JSON-ish list and a Python tuple repr, as
    well as lists that already arrived as Python objects. Embedded newlines are tolerated. There is
    deliberately no raw-string fallback: a scalar in a sequence column is a defect in the
    spreadsheet, not a one-element list.

    Parameters
    ----------
    value : Any
        A normalized cell value.

    Returns
    -------
    list | None
        The parsed list, or None if the value does not represent a sequence.
    """
    if isinstance(value, list | tuple):
        return list(value)
    if not isinstance(value, str):
        return None
    text = value.replace("\xa0", " ").strip()
    for parsed in (_try_literal_eval(text), _try_yaml_load(text)):
        if isinstance(parsed, list | tuple):
            return list(parsed)
    return None


def _coerce_number(value: Any) -> Any:
    """Coerce numeric-looking text to an int or float, leaving everything else alone.

    Parameters
    ----------
    value : Any
        A normalized cell value.

    Returns
    -------
    Any
        An int or float when the value looks numeric, otherwise the value unchanged.
    """
    if isinstance(value, bool) or not isinstance(value, str):
        return value
    parsed = _try_literal_eval(value)
    if isinstance(parsed, int | float) and not isinstance(parsed, bool) and not _is_nan(parsed):
        return parsed
    return value


def _collapse_newlines(text: str) -> str:
    """Collapse internal newlines and repeated whitespace in wrapped prose to single spaces.

    Parameters
    ----------
    text : str
        The prose to collapse.

    Returns
    -------
    str
        The same prose on one line.
    """
    return re.sub(r"\s+", " ", text).strip()


@dataclass(frozen=True)
class _Section:
    """The row bounds of one table on a sheet.

    Attributes
    ----------
    key: str
        Section key, one of the keys of :data:`SPREADSHEET_MARKERS`.
    marker_row: int
        Row holding the marker text in column A.
    header_row: int
        Row holding the column headers, always ``marker_row + 1``.
    first_data_row: int
        First row that may hold data.
    last_data_row: int
        Last row that may hold data, inclusive.
    """

    key: str
    marker_row: int
    header_row: int
    first_data_row: int
    last_data_row: int


def _find_sections(worksheet: Worksheet, product_id: DataProductIdentifier, source_name: str) -> dict[str, _Section]:
    """Locate the three tables on a sheet by scanning column A for anchored marker rows.

    Row positions vary from tab to tab, so nothing is keyed to a row number. A column A cell counts
    as a section marker only when the row immediately below it holds the matching header text; that
    anchoring is what distinguishes a real lowercase ``variables`` marker from a banner row that
    happens to use the same word.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet to scan.
    product_id : DataProductIdentifier
        The product being built, used only in error messages.
    source_name : str
        Display name of the source workbook, used only in error messages.

    Returns
    -------
    dict[str, _Section]
        Section key to its row bounds, for all three sections.

    Raises
    ------
    MalformedSheetError
        If a marker is missing, unanchored, or declared more than once.
    """
    expected = {
        key: (_normalize_marker(marker), _normalize_marker(header))
        for key, (marker, header) in SPREADSHEET_MARKERS.items()
    }
    found: dict[str, list[int]] = {key: [] for key in SPREADSHEET_MARKERS}
    max_row = worksheet.max_row or 0

    for row in range(1, max_row + 1):
        text = _normalize_marker(worksheet.cell(row=row, column=1).value)
        if not text:
            continue
        for key, (marker_text, header_text) in expected.items():
            if text == marker_text and _normalize_marker(worksheet.cell(row=row + 1, column=1).value) == header_text:
                found[key].append(row)

    missing = [key for key, rows in found.items() if not rows]
    if missing:
        details = "; ".join(
            f"no row in column A reads '{SPREADSHEET_MARKERS[key][0]}' with '{SPREADSHEET_MARKERS[key][1]}' "
            f"in the row directly beneath it"
            for key in missing
        )
        raise MalformedSheetError(
            f"Cannot build a product definition for {product_id} from sheet '{worksheet.title}' of {source_name}. "
            f"The sheet is missing {len(missing)} of its 3 required tables: {details}. Add the marker row and its "
            "header row to the sheet and re-run. No output file was written."
        )

    duplicated = {key: rows for key, rows in found.items() if len(rows) > 1}
    if duplicated:
        details = "; ".join(
            f"'{SPREADSHEET_MARKERS[key][0]}' appears at rows {', '.join(str(row) for row in rows)}"
            for key, rows in duplicated.items()
        )
        raise MalformedSheetError(
            f"Cannot build a product definition for {product_id} from sheet '{worksheet.title}' of {source_name}. "
            f"A section marker is declared more than once: {details}. Each table may appear only once per sheet. "
            "No output file was written."
        )

    marker_rows = sorted(rows[0] for rows in found.values())
    sections: dict[str, _Section] = {}
    for key, rows in found.items():
        marker_row = rows[0]
        later = [row for row in marker_rows if row > marker_row]
        last_data_row = (later[0] - 1) if later else max_row
        sections[key] = _Section(
            key=key,
            marker_row=marker_row,
            header_row=marker_row + 1,
            first_data_row=marker_row + 2,
            last_data_row=last_data_row,
        )
    return sections


def _lookup_column(header: str, allowed: dict[str, ColumnSpec]) -> ColumnSpec | None:
    """Resolve a header cell to a column specification.

    Parameters
    ----------
    header : str
        Normalized header text.
    allowed : dict[str, ColumnSpec]
        The allowlist for the table this header belongs to.

    Returns
    -------
    ColumnSpec | None
        The matching specification, or None if the header is not recognized.
    """
    if header in allowed:
        return allowed[header]
    if header.startswith(ENCODING_PREFIX):
        suffix = header[len(ENCODING_PREFIX) :].strip()
        if suffix:
            return ColumnSpec(suffix, ColumnTarget.ENCODING, ColumnKind.TEXT)
    return None


def _build_column_map(
    worksheet: Worksheet,
    header_row: int,
    allowed: dict[str, ColumnSpec],
    problems: _ProblemLog,
) -> dict[int, ColumnSpec]:
    """Map column indices to column specifications by reading a header row.

    Columns are always mapped by header text, never by index, because the number and order of
    columns varies between tabs. An unrecognized header is a problem rather than something to
    ignore, so that adding a column to the workbook forces a deliberate code update here instead of
    silently dropping the new field.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet being parsed.
    header_row : int
        Row holding the headers.
    allowed : dict[str, ColumnSpec]
        The allowlist for this table.
    problems : _ProblemLog
        Accumulator for problems found.

    Returns
    -------
    dict[int, ColumnSpec]
        1-based column index to specification, for every recognized column.
    """
    column_map: dict[int, ColumnSpec] = {}
    known = sorted(allowed)
    for column in range(1, (worksheet.max_column or 0) + 1):
        header = _normalize_marker(worksheet.cell(row=header_row, column=column).value)
        if not header:
            continue
        spec = _lookup_column(header, allowed)
        if spec is None:
            raw = worksheet.cell(row=header_row, column=column).value
            suggestion = difflib.get_close_matches(header, known, n=1, cutoff=0.6)
            hint = f" Did you mean '{suggestion[0]}'?" if suggestion else ""
            problems.add(
                header_row,
                column,
                f"'{raw}' is not a recognized column heading.{hint} Recognized headings for this table are: "
                f"{', '.join(known)}. If this column is genuinely new, it has to be added to the converter's "
                "column allowlist before it can be used.",
            )
            continue
        column_map[column] = spec
    return column_map


def _read_cell(
    worksheet: Worksheet,
    row: int,
    column: int,
    spec: ColumnSpec,
    problems: _ProblemLog,
    *,
    label: str,
) -> Any:
    """Read one cell and interpret it according to its column specification.

    Problems are recorded rather than raised so that a whole sheet's worth of defects can be
    reported at once.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet being parsed.
    row : int
        1-based row index.
    column : int
        1-based column index.
    spec : ColumnSpec
        How to interpret this column.
    problems : _ProblemLog
        Accumulator for problems found.
    label : str
        The spreadsheet-facing name of this column, used in problem messages.

    Returns
    -------
    Any
        :data:`EMPTY`, :data:`NOT_APPLICABLE`, or the interpreted value. Returns :data:`EMPTY` when
        a problem was recorded, so that parsing can continue.
    """
    value = _normalize_cell(worksheet.cell(row=row, column=column).value)
    if isinstance(value, _Sentinel):
        return value

    if _is_nan(value):
        problems.add(
            row,
            column,
            f"'{label}' is NaN, which cannot be written to a NetCDF file. Choose a sentinel value representable "
            "in the variable's dtype, for example -999 for an integer type or -9999.0 for float32.",
        )
        return EMPTY

    if spec.kind is ColumnKind.SEQUENCE:
        parsed = _parse_sequence(value)
        if parsed is None:
            problems.add(
                row,
                column,
                f"'{label}' must be a list, for example [\"RADIOMETER_TIME\"]. Found the scalar value "
                f"'{value}'. Wrap it in square brackets, even if there is only one entry.",
            )
            return EMPTY
        for element in parsed:
            if _is_nan(element):
                problems.add(row, column, f"'{label}' contains NaN, which cannot be written to a NetCDF file.")
                return EMPTY
        return parsed

    if spec.kind is ColumnKind.SCALAR:
        return _coerce_number(value)

    text = str(value)
    return _collapse_newlines(text) if spec.collapse_newlines else text


def _check_upper(name: str, row: int, column: int, kind: str, problems: _ProblemLog) -> None:
    """Record a problem if an identifier that must be uppercase is not.

    Mirrors the check already performed by ``DataProductIdentifier.__new__`` so that the two read
    the same way to a developer who hits both.

    Parameters
    ----------
    name : str
        The identifier to check.
    row : int
        1-based row index of the cell holding it.
    column : int
        1-based column index of the cell holding it.
    kind : str
        Spreadsheet-facing description, e.g. ``"Coordinate name"`` or ``"Dimension name"``.
    problems : _ProblemLog
        Accumulator for problems found.
    """
    if name != name.upper():
        problems.add(
            row,
            column,
            f"{kind} '{name}' is not uppercase. Coordinate and dimension names must be uppercase. "
            f"Expected '{name.upper()}'.",
        )


def _parse_metadata_section(
    worksheet: Worksheet,
    section: _Section,
    product_id: DataProductIdentifier,
    problems: _ProblemLog,
) -> dict[str, Any]:
    """Parse the product-specific file-level metadata table into product attributes.

    ``ProductID`` and ``algorithm_version`` are injected first, in that order, because the product
    definition model requires a statically defined ``ProductID`` and the sample YAML carries an
    explicit null ``algorithm_version``. A sheet that declares either of those itself does not
    produce a duplicate key. A sheet-declared ``ProductID`` must match the product being generated.

    Blank values become explicit nulls, marking attributes the algorithm supplies at write time.
    ``N/A`` values drop the attribute entirely.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet being parsed.
    section : _Section
        Row bounds of the metadata table.
    product_id : DataProductIdentifier
        The product being generated.
    problems : _ProblemLog
        Accumulator for problems found.

    Returns
    -------
    dict[str, Any]
        Product-level attributes in emission order.
    """
    column_map = _build_column_map(worksheet, section.header_row, METADATA_COLUMNS, problems)
    name_columns = [column for column, spec in column_map.items() if spec.target is ColumnTarget.NAME]
    value_columns = [column for column, spec in column_map.items() if spec.target is ColumnTarget.ATTRIBUTE]

    if not name_columns or not value_columns:
        problems.add(
            section.header_row,
            1,
            "The file-level metadata table needs an 'Attribute' column and a "
            "'Value (leave blank if calculated by the algorithm)' column. One or both are missing.",
        )
        return {"ProductID": str(product_id), "algorithm_version": None}

    name_column = name_columns[0]
    value_column = value_columns[0]
    value_spec = column_map[value_column]

    attributes: dict[str, Any] = {"ProductID": str(product_id), "algorithm_version": None}
    seen: dict[str, int] = {}

    for row in range(section.first_data_row, section.last_data_row + 1):
        raw_name = _normalize_cell(worksheet.cell(row=row, column=name_column).value)
        if isinstance(raw_name, _Sentinel):
            continue
        name = str(raw_name)

        if _normalize_marker(name) in RESERVED_BANNER_WORDS:
            problems.add(
                row,
                name_column,
                f"'{name}' can't be used as an attribute name because it's a YAML section marker. "
                "Rename it to something else.",
            )
            continue

        if name in seen:
            problems.add(
                row,
                name_column,
                f"Attribute '{name}' is already declared at "
                f"{problems.cell_reference(seen[name], name_column)}. Each attribute may appear only once. "
                "Delete the duplicate row.",
            )
            continue
        seen[name] = row

        value = _read_cell(worksheet, row, value_column, value_spec, problems, label=name)

        if name.lower() == "productid" and name != "ProductID":
            mismatch = ""
            if not isinstance(value, _Sentinel) and str(value) != str(product_id):
                mismatch = f" Note also that its value '{value}' does not match the product being generated, "
                mismatch += f"'{product_id}'."
            problems.add(
                row,
                name_column,
                f"'{name}' is not a recognized file-level attribute. Did you mean 'ProductID'?{mismatch}",
            )
            continue

        if name == "ProductID":
            if isinstance(value, _Sentinel) or str(value) != str(product_id):
                shown = "blank" if value is EMPTY else f"'{value}'"
                problems.add(
                    row,
                    value_column,
                    f"ProductID is {shown} but this sheet builds the product '{product_id}'. Set this cell to "
                    f"'{product_id}' or delete the row; the converter fills ProductID in either way.",
                )
            continue

        if value is NOT_APPLICABLE:
            continue
        attributes[name] = None if value is EMPTY else value

    return attributes


def _parse_variable_section(
    worksheet: Worksheet,
    section: _Section,
    problems: _ProblemLog,
    *,
    is_coordinate: bool,
    coordinate_cells: dict[str, tuple[int, int, str]] | None = None,
) -> tuple[dict[str, dict[str, Any]], dict[str, str], dict[str, tuple[int, int, str]]]:
    """Parse a Coordinates or Variables table into variable definitions.

    The two tables have the same row shape and go through the same code path; only the header of
    the name column and a handful of validation rules differ.

    Parameters
    ----------
    worksheet : Worksheet
        The sheet being parsed.
    section : _Section
        Row bounds of the table.
    problems : _ProblemLog
        Accumulator for problems found.
    is_coordinate : bool
        True when parsing the Coordinates table. Coordinate names must be uppercase; variable names
        are deliberately unconstrained.
    coordinate_cells : dict[str, tuple[int, int, str]] | None
        Coordinate names already parsed, keyed by lowercased name, mapping to the row, column, and
        original spelling of the cell that declared them. Supplied when parsing the Variables table
        so that a variable redeclaring a coordinate can be rejected.

    Returns
    -------
    tuple[dict[str, dict[str, Any]], dict[str, str], dict[str, tuple[int, int, str]]]
        The variable definitions in sheet order, the banner text preceding each variable, and the
        name-to-cell index for the names declared here.
    """
    allowed_names = coordinate_cells or {}
    column_map = _build_column_map(worksheet, section.header_row, VARIABLE_COLUMNS, problems)
    name_columns = [column for column, spec in column_map.items() if spec.target is ColumnTarget.NAME]
    data_columns = {
        column: spec
        for column, spec in sorted(column_map.items())
        if spec.target not in (ColumnTarget.NAME, ColumnTarget.IGNORE)
    }

    if not name_columns:
        problems.add(
            section.header_row,
            1,
            f"The {SPREADSHEET_MARKERS[section.key][0]} table has no "
            f"'{SPREADSHEET_MARKERS[section.key][1]}' column. Add that heading to the header row.",
        )
        return {}, {}, {}

    name_column = name_columns[0]
    definitions: dict[str, dict[str, Any]] = {}
    banners: dict[str, str] = {}
    declared_here: dict[str, tuple[int, int, str]] = {}
    pending_banner: str | None = None
    problems_before_rows = len(problems.problems)

    for row in range(section.first_data_row, section.last_data_row + 1):
        raw_name = _normalize_cell(worksheet.cell(row=row, column=name_column).value)
        others_empty = all(
            _normalize_cell(worksheet.cell(row=row, column=column).value) is EMPTY for column in data_columns
        )

        if isinstance(raw_name, _Sentinel):
            if not others_empty:
                problems.add(
                    row,
                    name_column,
                    f"This row has values but no {SPREADSHEET_MARKERS[section.key][1]}. Give it a name or delete "
                    "the row.",
                )
            continue

        name = str(raw_name)

        if others_empty:
            if _normalize_marker(name) in RESERVED_BANNER_WORDS:
                problems.add(
                    row,
                    name_column,
                    f"'{name}' can't be used as a banner row because it's a YAML section marker. Rename this "
                    f"banner to something else, for example '{worksheet.title} {name}'.",
                )
                continue
            pending_banner = name
            continue

        lowered = name.lower()
        if lowered in declared_here:
            previous_row, previous_column, previous_name = declared_here[lowered]
            problems.add(
                row,
                name_column,
                f"'{name}' is already declared at {problems.cell_reference(previous_row, previous_column)} "
                f"('{previous_name}'). Each name may appear only once in this table. Delete the duplicate row.",
            )
            continue
        if lowered in allowed_names:
            previous_row, previous_column, previous_name = allowed_names[lowered]
            problems.add(
                row,
                name_column,
                f"'{name}' is already declared as a coordinate at "
                f"{problems.cell_reference(previous_row, previous_column)} ('{previous_name}'). A name can be a "
                "coordinate or a variable, not both. Delete this row from the Variables table.",
            )
            continue
        declared_here[lowered] = (row, name_column, name)

        if is_coordinate:
            _check_upper(name, row, name_column, "Coordinate name", problems)

        definition: dict[str, Any] = {}
        attributes: dict[str, Any] = {}
        encoding: dict[str, Any] = {}
        already_reported: set[str] = set()

        for column, spec in data_columns.items():
            label = spec.key or _normalize_marker(worksheet.cell(row=section.header_row, column=column).value)
            problems_before = len(problems.problems)
            value = _read_cell(worksheet, row, column, spec, problems, label=label)

            if spec.target is ColumnTarget.TOP_LEVEL and spec.key == "dimensions":
                if not isinstance(value, _Sentinel):
                    dimensions = _validate_dimensions(value, row, column, problems)
                    if dimensions is not None:
                        definition["dimensions"] = dimensions
                if len(problems.problems) > problems_before:
                    already_reported.add(spec.key)
                continue

            if len(problems.problems) > problems_before:
                already_reported.add(spec.key)

            if isinstance(value, _Sentinel):
                # Both blank and N/A drop a variable-level key. Emitting a null here would silently
                # make the attribute a required dynamic attribute on every variable, which is not
                # what a blank cell in this table means.
                continue

            if spec.key == "valid_range" and isinstance(value, list) and not value:
                # An empty valid_range parses fine; it just means the attribute is not written.
                continue

            if spec.target is ColumnTarget.TOP_LEVEL:
                definition[spec.key] = value
            elif spec.target is ColumnTarget.ENCODING:
                encoding[spec.key] = value
            else:
                attributes[spec.key] = value

        for required in ("dtype", "dimensions"):
            # A key whose own cell already produced a problem is not reported a second time here;
            # one defect should read as one problem.
            if required not in definition and required not in already_reported:
                problems.add(
                    row,
                    name_column,
                    f"'{name}' has no {required}. Every coordinate and variable needs both a dtype and a "
                    "dimensions list.",
                )

        ordered: dict[str, Any] = {}
        for key in ("dtype", "dimensions"):
            if key in definition:
                ordered[key] = definition[key]
        ordered.update({key: value for key, value in definition.items() if key not in ordered})
        if attributes:
            ordered["attributes"] = attributes
        if encoding:
            ordered["encoding"] = encoding

        definitions[name] = ordered
        if pending_banner is not None:
            banners[name] = pending_banner
            pending_banner = None

    if not definitions and len(problems.problems) == problems_before_rows:
        # A table with banners but no rows means the tab has not been filled in yet. Emitting a
        # definition with no variables would produce a file that looks finished and is not. If rows
        # were present but every one of them was rejected, those rejections already say so, and
        # adding this on top would send a developer looking for a table that is right in front of
        # them.
        problems.add(
            section.header_row,
            name_column,
            f"The {SPREADSHEET_MARKERS[section.key][0]} table has no rows. Fill in at least one "
            f"{SPREADSHEET_MARKERS[section.key][1]} before generating a product definition from this sheet.",
        )

    return definitions, banners, declared_here


def _validate_dimensions(value: Any, row: int, column: int, problems: _ProblemLog) -> list[str] | None:
    """Validate a parsed dimensions list and return it as a list of uppercase names.

    An empty dimensions list is an error. This rule is deliberately not shared with
    ``valid_range``, where an empty list simply means the attribute is not written; a variable with
    no dimensions is not something the workbook is trying to express.

    Parameters
    ----------
    value : Any
        The value returned for the dimensions cell.
    row : int
        1-based row index of the cell.
    column : int
        1-based column index of the cell.
    problems : _ProblemLog
        Accumulator for problems found.

    Returns
    -------
    list[str] | None
        The validated dimension names, or None when a problem was recorded.
    """
    if not isinstance(value, list):
        return None
    if not value:
        problems.add(
            row,
            column,
            "'dimensions' is an empty list. Every coordinate and variable needs at least one dimension, for "
            'example ["RADIOMETER_TIME"].',
        )
        return None

    names: list[str] = []
    for element in value:
        if not isinstance(element, str):
            problems.add(
                row,
                column,
                f"'dimensions' contains {element!r}, which is not a dimension name. Dimensions must be a list "
                'of quoted names, for example ["RADIOMETER_TIME"].',
            )
            return None
        name = element.strip()
        _check_upper(name, row, column, "Dimension name", problems)
        names.append(name)
    return names


def find_product_sheets(workbook: Workbook) -> dict[str, str]:
    """Pair every tab name in a workbook with the ``Algorithm:`` value it declares.

    This is an audit helper, used to build useful error messages and to survey a workbook. It is
    not part of resolving a single request: under the tab-name-equals-Algorithm-equals-product-ID
    rule, resolution is a direct lookup.

    Parameters
    ----------
    workbook : Workbook
        An open workbook.

    Returns
    -------
    dict[str, str]
        Tab name to the value of its ``B1`` cell, empty string where ``B1`` is blank.
    """
    sheets: dict[str, str] = {}
    for sheet_name in workbook.sheetnames:
        declared = _normalize_cell(workbook[sheet_name][ALGORITHM_CELL].value)
        sheets[sheet_name] = "" if isinstance(declared, _Sentinel) else str(declared)
    return sheets


def resolve_sheet_name(workbook: Workbook, product_id: DataProductIdentifier) -> str:
    """Find the tab that defines a data product, checking that the tab agrees with itself.

    The tab name and the ``Algorithm:`` value in ``B1`` must be identical, and that shared string
    must be a valid data product identifier. Resolution is therefore a direct lookup, with ``B1``
    checked as a consistency assertion rather than used as an alternate key.

    Parameters
    ----------
    workbook : Workbook
        An open workbook.
    product_id : DataProductIdentifier
        The product being requested.

    Returns
    -------
    str
        The name of the tab defining that product.

    Raises
    ------
    ProductIdNotFoundError
        If no tab is named for the product.
    MalformedSheetError
        If the tab exists but its ``Algorithm:`` cell names something else.
    """
    sheets = find_product_sheets(workbook)
    wanted = str(product_id)

    if wanted not in sheets:
        near = [name for name, declared in sheets.items() if declared == wanted]
        near += difflib.get_close_matches(wanted, list(sheets), n=3, cutoff=0.6)
        deduped = list(dict.fromkeys(near))
        hint = ""
        if deduped:
            listed = ", ".join(f"'{name}'" for name in deduped)
            hint = (
                f" The workbook does contain {listed}, but a tab must be named exactly for the product it "
                "defines. Rename the tab, or ask for a product that has its own tab."
            )
        raise ProductIdNotFoundError(
            f"'{wanted}' is a valid data product identifier but the workbook has no tab named '{wanted}'.{hint} "
            f"Tabs in this workbook: {', '.join(sheets)}. No output file was written."
        )

    declared = sheets[wanted]
    if declared != wanted:
        shown = f"'{declared}'" if declared else "blank"
        raise MalformedSheetError(
            f"{wanted}!{ALGORITHM_CELL} declares {shown} but the tab is named '{wanted}'. The tab name and the "
            "Algorithm cell must match exactly, and both must be a valid data product identifier. No output file "
            "was written."
        )
    return wanted


def parse_product_definition_sheet(
    workbook: Workbook,
    sheet_name: str,
    product_id: DataProductIdentifier,
    source_name: str = "the workbook",
) -> ParsedSheet:
    """Parse one tab of the workbook into a product definition.

    Every problem found anywhere on the sheet is collected and reported together, so that a
    developer can correct the spreadsheet in one pass. The workbook is never modified.

    Parameters
    ----------
    workbook : Workbook
        An open workbook.
    sheet_name : str
        The tab to parse.
    product_id : DataProductIdentifier
        The product being generated.
    source_name : str
        Display name of the source workbook, used in error messages.

    Returns
    -------
    ParsedSheet
        The definition, plus the banner text that grouped the variables.

    Raises
    ------
    MalformedSheetError
        If the sheet's structure or contents cannot be interpreted.
    """
    worksheet = workbook[sheet_name]
    problems = _ProblemLog(sheet_name)
    sections = _find_sections(worksheet, product_id, source_name)

    attributes = _parse_metadata_section(worksheet, sections["attributes"], product_id, problems)
    coordinates, _, coordinate_cells = _parse_variable_section(
        worksheet, sections["coordinates"], problems, is_coordinate=True
    )
    variables, banners, _ = _parse_variable_section(
        worksheet, sections["variables"], problems, is_coordinate=False, coordinate_cells=coordinate_cells
    )

    problems.raise_if_any(product_id, source_name)

    definition = {"attributes": attributes, "coordinates": coordinates, "variables": variables}
    return ParsedSheet(definition=definition, banners=banners, sheet_name=sheet_name)


def _validate_definition(
    definition: dict[str, Any],
    product_id: DataProductIdentifier,
    sheet_name: str,
    source_name: str,
) -> None:
    """Confirm the assembled definition validates against the product definition model.

    A file on disk always means a definition that loads cleanly, so this is mandatory and
    hard-failing rather than something a flag can turn off. Validation failures are translated into
    the same message shape as parsing failures; no pydantic type names or tracebacks reach the user.

    Parameters
    ----------
    definition : dict[str, Any]
        The assembled definition.
    product_id : DataProductIdentifier
        The product being generated.
    sheet_name : str
        The tab the definition came from.
    source_name : str
        Display name of the source workbook.

    Raises
    ------
    MalformedSheetError
        If the definition does not validate.
    """
    try:
        LiberaDataProductDefinition.model_validate(definition)
        return
    except ValidationError as err:
        details = [
            (".".join(str(part) for part in error["loc"]), _clean_validation_message(str(error["msg"])))
            for error in err.errors()
        ]
    except (ValueError, TypeError) as err:
        details = [("", _clean_validation_message(str(err)))]

    rendered = "\n".join(f"  {location}: {message}" if location else f"  {message}" for location, message in details)
    count = len(details)
    plural = "problem" if count == 1 else "problems"
    raise MalformedSheetError(
        f"Cannot build a product definition for {product_id} from sheet '{sheet_name}' of {source_name}. "
        f"The definition assembled from that sheet is not a valid Libera product definition. Found {count} "
        f"{plural}:\n\n{rendered}\n\nCorrect these in the spreadsheet and re-run. No output file was written."
    )


def _clean_validation_message(message: str) -> str:
    """Strip pydantic's framing from a validation message so it reads as plain prose.

    Parameters
    ----------
    message : str
        The raw message from pydantic.

    Returns
    -------
    str
        The message without its ``Value error,`` style prefix, collapsed onto readable lines.
    """
    for prefix in ("Value error, ", "Assertion failed, ", "Type error, "):
        if message.startswith(prefix):
            message = message[len(prefix) :]
    return _collapse_newlines(message)


def spreadsheet_to_product_definition(
    spreadsheet_path: str | PathType,
    product_id: DataProductIdentifier,
) -> ParsedSheet:
    """Build a validated product definition from a tab of a data product specifications workbook.

    Opens the workbook read-only, resolves the tab for the requested product, parses it, and
    validates the result against the product definition model. Nothing is written anywhere.

    Parameters
    ----------
    spreadsheet_path : str | PathType
        Path to the workbook. May be a local path or a cloud path.
    product_id : DataProductIdentifier
        The product to build a definition for.

    Returns
    -------
    ParsedSheet
        The validated definition, plus the banner text that grouped the variables.

    Raises
    ------
    ProductIdNotFoundError
        If the workbook has no tab for the requested product.
    MalformedSheetError
        If the tab cannot be interpreted or the resulting definition does not validate.
    """
    path = cast(PathType, AnyPath(spreadsheet_path))
    source_name = path.name
    logger.info(f"Reading product definition for {product_id} from {path}")

    with path.open("rb") as spreadsheet_file:
        workbook = openpyxl.load_workbook(spreadsheet_file, data_only=True, read_only=False)
        try:
            sheet_name = resolve_sheet_name(workbook, product_id)
            parsed = parse_product_definition_sheet(workbook, sheet_name, product_id, source_name)
        finally:
            workbook.close()

    _validate_definition(parsed.definition, product_id, parsed.sheet_name, source_name)
    logger.info(
        f"Parsed {len(parsed.definition['coordinates'])} coordinates and "
        f"{len(parsed.definition['variables'])} variables from sheet '{parsed.sheet_name}'"
    )
    return parsed


class _ProductDefinitionDumper(yaml.SafeDumper):
    """A safe dumper that keeps short scalar lists inline.

    Mappings are emitted in block style, but a list of scalars is emitted as ``[a, b]`` rather than
    expanded over several lines. This matches the hand-written Libera product definitions, where
    ``dimensions`` and ``valid_range`` read as one-liners, so a generated file diffs cleanly against
    one a person wrote.
    """


def _represent_list(dumper: yaml.SafeDumper, data: list) -> yaml.Node:
    """Represent a list, using flow style when every element is a scalar.

    Strings inside a flow sequence are double quoted so that ``dimensions`` reads as
    ``["RADIOMETER_TIME"]``, matching the hand-written product definition files. That is purely
    cosmetic to YAML, but it keeps a generated file diffable against the one it replaces.

    Parameters
    ----------
    dumper : yaml.SafeDumper
        The dumper doing the work.
    data : list
        The list being represented.

    Returns
    -------
    yaml.Node
        The sequence node.
    """
    scalar_only = all(item is None or isinstance(item, str | int | float | bool) for item in data)
    node = dumper.represent_sequence("tag:yaml.org,2002:seq", data, flow_style=scalar_only)
    if scalar_only:
        for child in node.value:
            if child.tag == "tag:yaml.org,2002:str":
                child.style = '"'
    return node


def _represent_str(dumper: yaml.SafeDumper, data: str) -> yaml.Node:
    """Represent a string, using a literal block for multi-line prose.

    Bit-flag ``value_meaning`` cells use one line per flag and that structure is meaningful. Without
    this, PyYAML emits them as a single quoted scalar full of escaped newlines, which round-trips
    correctly but is unreadable to the person maintaining the file.

    Parameters
    ----------
    dumper : yaml.SafeDumper
        The dumper doing the work.
    data : str
        The string being represented.

    Returns
    -------
    yaml.Node
        The scalar node.
    """
    if "\n" in data:
        return dumper.represent_scalar("tag:yaml.org,2002:str", data, style="|")
    return dumper.represent_scalar("tag:yaml.org,2002:str", data)


_ProductDefinitionDumper.add_representer(list, _represent_list)
_ProductDefinitionDumper.add_representer(str, _represent_str)


def _dump_mapping(mapping: dict[str, Any]) -> str:
    """Dump a mapping to YAML text without sorting keys.

    Parameters
    ----------
    mapping : dict[str, Any]
        The mapping to dump.

    Returns
    -------
    str
        YAML text, or ``{}`` for an empty mapping.
    """
    if not mapping:
        return "{}"
    return yaml.dump(
        mapping,
        Dumper=_ProductDefinitionDumper,
        sort_keys=False,
        default_flow_style=False,
        width=120,
        allow_unicode=True,
    ).rstrip("\n")


def _banner_block(text: str) -> list[str]:
    """Render a banner row as the boxed comment style used in the hand-written definitions.

    Parameters
    ----------
    text : str
        The banner text.

    Returns
    -------
    list[str]
        Three comment lines, already indented for placement inside ``variables:``.
    """
    single = _collapse_newlines(text)
    rule = "#" * (len(single) + 4)
    return [f"  {rule}", f"  # {single} #", f"  {rule}"]


def render_product_definition_yaml(
    definition: dict[str, Any],
    banners: dict[str, str] | None = None,
    header_comment: str | None = None,
) -> str:
    """Render a product definition as YAML text, with banner comments and a provenance header.

    The document is assembled in chunks rather than dumped in one call, because ``safe_dump``
    cannot interleave the comment lines that the hand-written definitions use to group variables.

    Parameters
    ----------
    definition : dict[str, Any]
        The definition, with ``attributes``, ``coordinates``, and ``variables`` keys.
    banners : dict[str, str] | None
        Maps a variable name to the banner text emitted just above it.
    header_comment : str | None
        A provenance line emitted at the top of the file, without its leading ``#``.

    Returns
    -------
    str
        The complete YAML document.
    """
    banners = banners or {}
    lines: list[str] = []
    if header_comment:
        lines.extend([f"# {header_comment}", ""])

    for key in DEFINITION_KEY_ORDER:
        section = definition.get(key, {})
        if key != "variables":
            lines.append(f"{key}:")
            lines.append(textwrap.indent(_dump_mapping(section), "  "))
            lines.append("")
            continue

        lines.append(f"{key}:")
        if not section:
            lines[-1] = f"{key}: {{}}"
            lines.append("")
            continue
        for name, variable in section.items():
            if name in banners:
                lines.extend(_banner_block(banners[name]))
            lines.append(textwrap.indent(_dump_mapping({name: variable}), "  "))
        lines.append("")

    return "\n".join(lines).rstrip("\n") + "\n"


def write_product_definition_yaml(
    definition: dict[str, Any],
    output_path: str | PathType,
    banners: dict[str, str] | None = None,
    header_comment: str | None = None,
    overwrite: bool = False,
) -> PathType:
    """Write a product definition to a YAML file.

    Parameters
    ----------
    definition : dict[str, Any]
        The definition to write.
    output_path : str | PathType
        Where to write it.
    banners : dict[str, str] | None
        Maps a variable name to the banner text emitted just above it.
    header_comment : str | None
        A provenance line emitted at the top of the file, without its leading ``#``.
    overwrite : bool
        Whether to replace an existing file.

    Returns
    -------
    PathType
        The path written.

    Raises
    ------
    FileExistsError
        If the file exists and ``overwrite`` is False.
    """
    path = cast(PathType, AnyPath(output_path))
    if path.exists() and not overwrite:
        raise FileExistsError(
            f"{path} already exists. Pass --overwrite to replace it, or choose a different output path with "
            "--output-path. No output file was written."
        )
    document = render_product_definition_yaml(definition, banners=banners, header_comment=header_comment)
    with path.open("w") as output_file:
        output_file.write(document)
    logger.info(f"Wrote product definition to {path}")
    return path


def data_product_identifier(raw: str) -> DataProductIdentifier:
    """Convert a command line string into a data product identifier, with useful hinting on failure.

    Used as an argparse ``type`` converter rather than ``choices`` so that the error message can be
    shaped: an unrecognized value gets close matches first and the full list second, grouped by data
    level so it reads as something other than a wall of text.

    Parameters
    ----------
    raw : str
        The raw command line value.

    Returns
    -------
    DataProductIdentifier
        The matching identifier.

    Raises
    ------
    argparse.ArgumentTypeError
        If the value is not a valid data product identifier.
    """
    candidate = raw.strip().upper()
    try:
        return DataProductIdentifier(candidate)
    except ValueError:
        pass

    known = [str(item) for item in DataProductIdentifier]
    close = difflib.get_close_matches(candidate, known, n=5, cutoff=0.6)
    hint = f" Did you mean {', '.join(close)}?" if close else ""

    by_level: dict[str, list[str]] = {}
    for item in DataProductIdentifier:
        by_level.setdefault(str(item.level), []).append(str(item))
    grouped = "\n".join(f"  {level}: {', '.join(sorted(names))}" for level, names in sorted(by_level.items()))

    raise argparse.ArgumentTypeError(
        f"'{raw}' is not a valid data product identifier.{hint}\nValid identifiers, by data level:\n{grouped}"
    )


def product_definition_spreadsheet_cli_handler(parsed_args: argparse.Namespace) -> PathType:
    """Convert a spreadsheet tab into a product definition YAML file from command line arguments.

    Parameters
    ----------
    parsed_args : argparse.Namespace
        Namespace of parsed CLI arguments.

    Returns
    -------
    PathType
        The path of the YAML file written.
    """
    now = datetime.now(UTC).strftime("%Y%m%dT%H%M%S")
    configure_task_logging(
        f"product_definition_from_spreadsheet_{now}",
        limit_debug_loggers=["libera_utils"],
        console_log_level=logging.DEBUG if parsed_args.verbose else logging.INFO,
    )

    product_id = parsed_args.product_id
    spreadsheet_path = cast(PathType, AnyPath(parsed_args.spreadsheet))
    parsed = spreadsheet_to_product_definition(spreadsheet_path, product_id)

    output_path = parsed_args.output_path or f"{product_id}_product_definition.yml"
    header_comment = (
        f"Auto-generated by libera-utils {libera_utils_version()} from {spreadsheet_path.name}, "
        f'sheet "{parsed.sheet_name}". Do not edit by hand.'
    )
    return write_product_definition_yaml(
        parsed.definition,
        output_path,
        banners=parsed.banners,
        header_comment=header_comment,
        overwrite=parsed_args.overwrite,
    )
